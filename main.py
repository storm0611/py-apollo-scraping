from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
from dotenv import load_dotenv
import os
import undetected_chromedriver as uc
from math import ceil
import openpyxl
from database import my_sqlite

load_dotenv()

# pids = []
data = []
TIMEOUT = 10
filter_query = os.environ.get("FILTER_QUERY", "")
is_free = int(os.environ.get("IS_FREE", 1))

def export():
    global data
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add some data to the worksheet
    worksheet['A1'] = 'ID'
    worksheet['B1'] = 'Name'
    worksheet['C1'] = 'Personal LinkedIn'
    worksheet['D1'] = 'Email'
    worksheet['E1'] = 'Job Title'
    worksheet['F1'] = 'Person Apollo Link'
    worksheet['G1'] = 'Company Name'
    worksheet['H1'] = 'Company Apollo Link'
    worksheet['I1'] = 'Company Website'
    worksheet['J1'] = 'Company LinkedIn'
    worksheet['K1'] = 'Company Twitter'
    worksheet['L1'] = 'Company Facebook'
    worksheet['M1'] = 'Location'
    worksheet['N1'] = 'Employees'
    worksheet['O1'] = 'Industry'
    worksheet['P1'] = 'Keywords'

    for i in range(len(data)):
        item = data[i]
        worksheet.cell(row=i + 2, column=1).value = item["person_id"]
        worksheet.cell(row=i + 2, column=2).value = item["name"]
        worksheet.cell(row=i + 2, column=3).value = '=HYPERLINK("{}")'.format(item["person_linkedin"])
        worksheet.cell(row=i + 2, column=4).value = item["email"]
        worksheet.cell(row=i + 2, column=5).value = item["job_title"]
        worksheet.cell(row=i + 2, column=6).value = '=HYPERLINK("{}")'.format(item["person_link"])
        worksheet.cell(row=i + 2, column=7).value = item["company_name"]
        worksheet.cell(row=i + 2, column=8).value = '=HYPERLINK("{}")'.format(item["company_link"])
        worksheet.cell(row=i + 2, column=9).value = '=HYPERLINK("{}")'.format(item["company_website"])
        worksheet.cell(row=i + 2, column=10).value = '=HYPERLINK("{}")'.format(item["company_linkedin"])
        worksheet.cell(row=i + 2, column=11).value = '=HYPERLINK("{}")'.format(item["company_twitter"])
        worksheet.cell(row=i + 2, column=12).value = '=HYPERLINK("{}")'.format(item["company_facebook"])
        worksheet.cell(row=i + 2, column=13).value = item["location"]
        worksheet.cell(row=i + 2, column=14).value = item["employees"]
        worksheet.cell(row=i + 2, column=15).value = item["industry"]
        worksheet.cell(row=i + 2, column=16).value = ",".join(item["keywords"])

    # Save the workbook
    try:
        # Set the filename for the workbook
        filename = str(int(datetime.now().timestamp()))+'.xlsx'
        # Set the directory to save the file in
        save_dir = "./csv"
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
        # Create the full path to the file
        filepath = os.path.join(save_dir, filename)
        # Save the workbook to the file
        workbook.save(filepath)
        workbook.close()
    except Exception as err:
        print(str(err))

def export_one(data):

    filename = 'data.xlsx'
    save_dir = './csv'
    if not os.path.exists(save_dir):
        os.mkdir(save_dir)
    filepath = os.path.join(save_dir, filename)
    if os.path.exists(filepath):
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'ID'
        worksheet['B1'] = 'Name'
        worksheet['C1'] = 'Personal LinkedIn'
        worksheet['D1'] = 'Email'
        worksheet['E1'] = 'Job Title'
        worksheet['F1'] = 'Person Apollo Link'
        worksheet['G1'] = 'Company Name'
        worksheet['H1'] = 'Company Apollo Link'
        worksheet['I1'] = 'Company Website'
        worksheet['J1'] = 'Company LinkedIn'
        worksheet['K1'] = 'Company Twitter'
        worksheet['L1'] = 'Company Facebook'
        worksheet['M1'] = 'Location'
        worksheet['N1'] = 'Employees'
        worksheet['O1'] = 'Industry'
        worksheet['P1'] = 'Keywords'
        worksheet['Q1'] = 'Phone Number'

    start_row = worksheet.max_row + 1

    worksheet.cell(row=start_row, column=1).value = data["person_id"]
    worksheet.cell(row=start_row, column=2).value = data["name"]
    worksheet.cell(row=start_row, column=3).value = '=HYPERLINK("{}")'.format(data["person_linkedin"])
    worksheet.cell(row=start_row, column=4).value = data["email"]
    worksheet.cell(row=start_row, column=5).value = data["job_title"]
    worksheet.cell(row=start_row, column=6).value = '=HYPERLINK("{}")'.format(data["person_link"])
    worksheet.cell(row=start_row, column=7).value = data["company_name"]
    worksheet.cell(row=start_row, column=8).value = '=HYPERLINK("{}")'.format(data["company_link"])
    worksheet.cell(row=start_row, column=9).value = '=HYPERLINK("{}")'.format(data["company_website"])
    worksheet.cell(row=start_row, column=10).value = '=HYPERLINK("{}")'.format(data["company_linkedin"])
    worksheet.cell(row=start_row, column=11).value = '=HYPERLINK("{}")'.format(data["company_twitter"])
    worksheet.cell(row=start_row, column=12).value = '=HYPERLINK("{}")'.format(data["company_facebook"])
    worksheet.cell(row=start_row, column=13).value = data["location"]
    worksheet.cell(row=start_row, column=14).value = data["employees"]
    worksheet.cell(row=start_row, column=15).value = data["industry"]
    worksheet.cell(row=start_row, column=16).value = ",".join(data["keywords"])
    worksheet.cell(row=start_row, column=17).value = data["phone_number"]

    # Save the workbook
    workbook.save(filepath)
    workbook.close()

def signup(driver: uc.Chrome):
    driver.get("https://www.apollo.io/sign-up")
    action = ActionChains(driver)
    element = WebDriverWait(driver, TIMEOUT).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'input[type="text"]'))
    )
    element.clear()
    action.send_keys_to_element(element, "luckmanguyen85@gmail.com")
    element = WebDriverWait(driver, TIMEOUT).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'input.PrivateSwitchBase-input'))
    )
    element.clear()
    action.move_to_element(element).click().perform()

def login(driver: uc.Chrome):
    driver.get("https://app.apollo.io/#/login")
    action = ActionChains(driver)
    element = WebDriverWait(driver, TIMEOUT).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'input[name="email"]'))
    )
    element.clear()
    action.send_keys_to_element(element, os.environ.get("EMAIL")).perform()
    element = WebDriverWait(driver, TIMEOUT).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'input[name="password"]'))
    )
    element.clear()
    action.send_keys_to_element(element, os.environ.get("PASSWORD")).perform()
    element = WebDriverWait(driver, TIMEOUT).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'button[type="submit"]'))
    )
    element.click()
    WebDriverWait(driver, TIMEOUT).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, '#side-nav-search'))
    )
    

def filter(driver: uc.Chrome, query: str):
    driver.get("https://app.apollo.io/#/people")
    WebDriverWait(driver, TIMEOUT).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, '.pipeline-tabs > a:first-child'))
    )
    url = f"{driver.current_url}&{query}" if "?" in driver.current_url else f"{driver.current_url}?{query}"
    driver.get(url)
    action = ActionChains(driver)
    try:
        elements = WebDriverWait(driver, TIMEOUT).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.apolloio-css-vars-reset .zp-modal-content button'))
        )
        for element in elements:
            if "Skip" in element.find_element(By.CSS_SELECTOR, 'div[data-elem="button-label"]').text:
                print("Skip button click on dialog")
                element.click()
    except:
        print("Not dialog")
        pass
    element = WebDriverWait(driver, TIMEOUT).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, '.pipeline-tabs > a:first-child'))
    )
    action.move_to_element(element).click()
    element = WebDriverWait(driver, TIMEOUT).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'div.finder-results-list-panel-content > div > div > div > div > div:last-child > div > div > span'))
    )
    count = int(element.text.split("of ")[1].replace(",", ""))
    if is_free:
        max_page_num = int(ceil(count / 25)) if ceil(count / 25) <= 5 else 5
    else:
        max_page_num = int(ceil(count / 25))
    print("max_page_num = ", max_page_num)
    skip = 0
    i = 0
    while i <= max_page_num:
        i += 1
        driver.get(f"{url}&page={i}")
        element = WebDriverWait(driver, TIMEOUT).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.pipeline-tabs > a:first-child'))
        )
        action.move_to_element(element).click()
        elements = WebDriverWait(driver, TIMEOUT).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.finder-results-list-panel-content table tbody tr'))
        )
        print(f"current page: {i}, count: {len(elements)}")
        for element in elements:
            skip = 0
            person_id = "Not Found"
            person_linkedin = "Not Found"
            person_link = "Not Found"
            name = "Not Found"
            title = "Not Found"
            company_name = "Not Found"
            company_link = "Not Found"
            company_website = "Not Found"
            company_linkedin = "Not Found"
            company_twitter = "Not Found"
            company_facebook = "Not Found"
            location = "Not Found"
            employees = "Not Found"
            industry = "Not Found"
            email = "Not Found"
            phone_number = "Not Found"
            keywords = []
            columns = element.find_elements(By.TAG_NAME, 'td')
            print(f"{len(columns)} Columns Found")
    # for j in range(len(columns)):
        # if j == 0:
            if not len(columns):
                continue
            for sub_col in columns[0].find_elements(By.TAG_NAME, 'a'):
                link = sub_col.get_attribute("href")
                if "linkedin.com" in link:
                    person_linkedin = link
                else:
                    person_link = link
                    name = sub_col.text
                    person_id = link.split("/")[-1].strip()
                    if len(my_sqlite.select(pid=person_id)):
                    # if person_id in pids:
                        skip = 1
                        break
                    my_sqlite.insert(pid=person_id)
                    # pids.append(person_id)
            if skip:
                break
        # elif j == 1:
            try:
                title = columns[1].find_element(By.CSS_SELECTOR, 'span.zp_Y6y8d').text
            except:
                pass
        # elif j == 2:
            try:
                sub_col = columns[2].find_element(By.CSS_SELECTOR, 'div.zp_J1j17 a')
                company_link = sub_col.get_attribute("href")
                company_name = sub_col.text
                for sub_col in columns[2].find_elements(By.CSS_SELECTOR, 'div.zp_I1ps2 a'):
                    link = sub_col.get_attribute("href")
                    if "linkedin.com" in link:
                        company_linkedin = link
                    elif "twitter.com" in link:
                        company_twitter = link
                    elif "facebook.com" in link:
                        company_facebook = link
                    else:
                        company_website = link
            except:
                pass
        # elif j == 3:
            # pass
        # elif j == 4:
            try:
                location = columns[4].find_element(By.CSS_SELECTOR, 'span.zp_Y6y8d').text
            except:
                pass
        # elif j == 5:
            try:
                employees = columns[5].find_element(By.CSS_SELECTOR, 'span.zp_Y6y8d').text
            except:
                pass
        # elif j == 7:
            try:
                industry = columns[7].find_element(By.CSS_SELECTOR, 'div.zp_paOF8 > span').text.split(',')[0]
            except:
                pass
        # elif j == 8:
            try:
                sub_cols = columns[8].find_elements(By.CSS_SELECTOR, 'div.zp_HlgrG > span')
                if len(sub_cols):
                    for sub_col in sub_cols:
                        keyword = sub_col.text.replace(",", "")
                        if keyword != "":
                            keywords.append(keyword)
            except:
                pass
        # elif j == 6:
            try:
                element = columns[6].find_element(By.CSS_SELECTOR, 'a.zp-link')
                if "@" in element.text:
                    email = element.text
                else:
                    phone_number = element.text
                    button = columns[3].find_element(By.CSS_SELECTOR, 'button:first-child')
                    button.click()
                    time.sleep(3)
                    try:
                        element = WebDriverWait(driver, TIMEOUT).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, 'div.apolloio-css-vars-reset div.apolloio-css-vars-reset span.zp_t08Bv'))
                        )
                        email = element.text
                    except:
                        pass
            except:
                try:
                    button = columns[6].find_element(By.CSS_SELECTOR, 'button')
                    button.find_element(By.CSS_SELECTOR, 'div[data-elem="button-label"]').text
                    button.click()
                    time.sleep(3)
                    try:
                        element = columns[6].find_element(By.CSS_SELECTOR, 'a.zp-link')
                        if "@" in element.text:
                            email = element.text
                        else:
                            phone_number = element.text
                            button = columns[3].find_element(By.CSS_SELECTOR, 'button:first-child')
                            button.click()
                            time.sleep(3)
                            try:
                                element = WebDriverWait(driver, TIMEOUT).until(
                                    EC.presence_of_element_located((By.CSS_SELECTOR, 'div.apolloio-css-vars-reset div.apolloio-css-vars-reset span.zp_t08Bv'))
                                )
                                email = element.text
                            except:
                                pass
                    except:
                        pass
                except:
                    pass
            if not skip:
                export_one({
                    "person_id": person_id,
                    "name": name,
                    "person_linkedin": person_linkedin,
                    "email": email,
                    "job_title": title,
                    "person_link": person_link,
                    "company_name": company_name,
                    "company_link": company_link,
                    "company_website": company_website,
                    "company_linkedin": company_linkedin,
                    "company_twitter": company_twitter,
                    "company_facebook": company_facebook,
                    "location": location,
                    "employees": employees,
                    "industry": industry,
                    "keywords": keywords,
                    "phone_number": phone_number,
                })
                # data.append({
                #     "person_id": person_id,
                #     "name": name,
                #     "person_linkedin": person_linkedin,
                #     "email": email,
                #     "job_title": title,
                #     "person_link": person_link,
                #     "company_name": company_name,
                #     "company_link": company_link,
                #     "company_website": company_website,
                #     "company_linkedin": company_linkedin,
                #     "company_twitter": company_twitter,
                #     "company_facebook": company_facebook,
                #     "location": location,
                #     "employees": employees,
                #     "industry": industry,
                #     "keywords": keywords,
                # })
                print({
                    "person_id": person_id,
                    "name": name,
                    "person_linkedin": person_linkedin,
                    "email": email,
                    "job_title": title,
                    "person_link": person_link,
                    "company_name": company_name,
                    "company_link": company_link,
                    "company_website": company_website,
                    "company_linkedin": company_linkedin,
                    "company_twitter": company_twitter,
                    "company_facebook": company_facebook,
                    "location": location,
                    "employees": employees,
                    "industry": industry,
                    "keywords": keywords,
                    "phone_number": phone_number,
                })

if __name__ == "__main__":
    # while True:
    # try:
        options = webdriver.ChromeOptions()
        options.headless = False
        driver = uc.Chrome(options=options)
        driver.maximize_window()
        login(driver)
        filter(driver, filter_query)
        driver.quit()
    # except Exception as error:
        # print(error)
    # export()
    # time.sleep(3)
